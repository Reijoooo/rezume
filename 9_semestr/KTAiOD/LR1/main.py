import math
from sklearn.linear_model import LinearRegression
import pandas as pd
import openpyxl
import datetime
import numpy as np
import matplotlib.pyplot as plt
import statsmodels.api as sm
from matplotlib import pyplot
from statsmodels.tsa.seasonal import seasonal_decompose
from sklearn.linear_model import LogisticRegression
from scipy.optimize import minimize
from scipy.optimize import curve_fit

def model_function(t, k, a, b):
    return k / (1 + a * np.exp(b * t))

    # Функция потерь для оптимизации

def loss_function(params, t, y):
    k, a, b = params
    t_values = np.arange(len(t))  # создаем значения t как индексы
    y_pred = model_function(t_values, k, a, b)
    return np.mean((y - y_pred) ** 2)

def func1():
    initial_params = [150, 1, 1]  # начальные значения для k, a, b

    # Оптимизация параметров
    optimal_params = minimize(loss_function, initial_params, args=(t, y), method='L-BFGS-B').x

    # Разделение оптимальных параметров
    k_optimal, a_optimal, b_optimal = optimal_params

    # Генерация значений t для предсказаний
    t_values = np.arange(len(t))

    # Предсказание на основе оптимальных параметров
    y_pred_optimal = model_function(t_values, k_optimal, a_optimal, b_optimal)
    #print(y_pred_optimal)
    # Построение графика с исходными данными и моделью
    plt.scatter(t, y, color='blue', label='Исходные данные')
    plt.plot(t, y_pred_optimal, color='red', label='Модель 1')
    plt.xlabel('t')
    plt.ylabel('Метка класса')
    plt.legend()
    plt.show()

    # Вывод оптимальных параметров
    print('Оптимальные параметры:')
    print('k:', k_optimal)
    print('a:', a_optimal)
    print('b:', b_optimal)

    #Высчитываем максимальный и минимальный остаток 1 модели
    # residuals = [actual - forecast for actual, forecast in zip(y, y_pred_optimal)]
    # max_residual = max(residuals)
    # print("Максимальный остаток 1 модели:", max_residual)
    # min_residual = min(residuals)
    # print("Минимальный остаток 1 модели:", min_residual)
    # mean_residual = np.mean(residuals)
    # print("Средняя ошибка 1:", mean_residual)
    # std_residual = np.std(residuals)
    # print("Стандартное отклонение ошибки 1:", std_residual)
    # print()

if __name__ == '__main__':
    # Чтение файла Excel
    wb = openpyxl.load_workbook('date_list.xlsx')
    sheet = wb['Лист1']
    # Инициализация списков для данных
    y = []
    t = []
    # Чтение данных из разных столбцов
    for i in range(2, 193):
        y.append(sheet.cell(row=i, column=18).value) #Вариант 15
        t.append(sheet.cell(row=i, column=3).value) #Типа X числа до 192

    # Вычисляем x_train
    # 3 лага - взяты значения за 3 месяца
    lag = 3
    x_train = []
    y_train1 = []
    y_train2 = []
    for i in range(lag, len(t)):
        x_train.append(t[i - lag:i])
    x_train = np.array(x_train)


    wb.close()
    plt.grid(True)
    plt.plot(t, y)# 1 график
    ts = pd.Series(y, index=t)
    ts.dropna(inplace=True)
    result = sm.tsa.seasonal_decompose(ts, model='additive', period=12)
    # result.plot()# 2 график
    residuals = result.resid
    trend = result.trend
    result.plot()  # 2 график
    plt.plot(residuals, marker='', linestyle='-') #2.4 график
    plt.figure()

    #6 функция
    #Далее берем из таблицу 1 в методички функции 6,7
    # Берем сначал 1 функцию и прогоняем её
    # k = 150
    # a, b - кофф. которые мы должны вяснить
    #func1()
    # График функции y = k / (1 + a * e ** (b * t))
    k_function = 144  # Изменил переменную k
    a_function = 13  # Изменил переменную a
    b_function = -0.097  # Изменил переменную b

    # Рассчитываем значения функции для каждого t
    y_function = [k_function / (1 + a_function * math.e ** (b_function * t_val)) for t_val in t]

    #Уравниваем массивы для просчета следующих 3 шагов
    y_train1 = y_function
    min_length = min(len(x_train), len(y_train1))
    x_train = x_train[:min_length]
    y_train1 = y_train1[:min_length]

    plt.figure()

    #Высчитываем максимальный и минимальный остаток 1 модели
    residuals = [actual - forecast for actual, forecast in zip(y, y_function)]
    min_residual = min(residuals)
    print("Минимальный остаток 1 модели:", min_residual)
    max_residual = max(residuals)
    print("Максимальный остаток 1 модели:", max_residual)
    mean_residual = np.mean(residuals)
    print("Средняя ошибка 1:", mean_residual)
    std_residual = np.std(residuals)
    print("Стандартное отклонение ошибки 1:", std_residual)
    mae = sum(abs(residual) for residual in residuals) / len(residuals)
    print("Средняя абсолютная ошибка (MAE) 1:", mae)
    mae_percentage = (sum((actual - forecast) / actual for actual, forecast in zip(y, y_function)) / len(y)) * 100
    print("Средняя ошибка в процентах 1:", mae_percentage, "%")
    mape = sum(abs((actual - forecast) / actual) * 100 for actual, forecast in zip(y, y_function)) / len(y)
    print("Средняя абсолютная процентная ошибка (MAPE) 1:", mape, "%")
    mse = sum((actual - forecast) ** 2 for actual, forecast in zip(y, y_function)) / len(y)
    print("Средняя квадратичная ошибка (MSE) 1:", mse)
    actual_mean = np.mean(y)
    sst = sum((actual - actual_mean) ** 2 for actual in y)
    sse = sum((actual - forecast) ** 2 for actual, forecast in zip(y, y_function))
    r_squared = 1 - (sse / sst)
    print("Коэффициент детерминации (R-squared) 1:", r_squared)
    #Прогноз на 3 шага вперёд 1 модель
    model = LinearRegression()
    model.fit(x_train, y_train1)  # y - входные данные, y_function - фактические значения
    future_input = [(193, 181,  140.54365437), (194, 182, 138.4325965435), (195, 183, 142.244325656)] # Данные для следующих 3 шагов
    forecast = model.predict(future_input)
    print("Прогноз на 3 шага вперед 1 модели:", forecast)
    print()
    plt.plot(t, y_function)  # 3 график
    plt.plot(trend)                # 3 график
    plt.grid(True)
    plt.show()

    #7 функция
    k1_function = 140.0020512962645  # Изменил переменную k
    a1_function = 0.04303915207389439  # Изменил переменную a
    b1_function = 0.9380820757016298  # Изменил переменную b

    f_t = k1_function * np.power(a1_function, np.power(b1_function, t))

    # Уравниваем массивы для просчета следующих 3 шагов
    y_train2 = f_t
    min_length = min(len(x_train), len(y_train2))
    x_train = x_train[:min_length]
    y_train2 = y_train2[:min_length]

    #Высчитываем максимальный и минимальный остаток 2 модели
    residuals2 = [actual - forecast for actual, forecast in zip(y, f_t)]
    min_residual2 = min(residuals2)
    print("Минимальный остаток 2 модели:", min_residual2)
    max_residual2 = max(residuals2)
    print("Максимальный остаток 2 модели:", max_residual2)
    mean_residual2 = np.mean(residuals2)
    print("Средняя ошибка 2:", mean_residual2)
    std_residual2 = np.std(residuals2)
    print("Стандартное отклонение ошибки 2:", std_residual2)
    mae2 = sum(abs(residual2) for residual2 in residuals2) / len(residuals2)
    print("Средняя абсолютная ошибка (MAE) 2:", mae2)
    mae_percentage2 = (sum((actual - forecast) / actual for actual, forecast in zip(y, f_t)) / len(y)) * 100
    print("Средняя ошибка в процентах 2:", mae_percentage2, "%")
    mape2 = sum(abs((actual - forecast) / actual) * 100 for actual, forecast in zip(y, f_t)) / len(y)
    print("Средняя абсолютная процентная ошибка (MAPE) 2:", mape2, "%")
    mse2 = sum((actual - forecast) ** 2 for actual, forecast in zip(y, f_t)) / len(y)
    print("Средняя квадратичная ошибка (MSE) 1:", mse2)
    actual_mean2 = np.mean(y)
    sst2 = sum((actual - actual_mean2) ** 2 for actual in y)
    sse2 = sum((actual - forecast) ** 2 for actual, forecast in zip(y, f_t))
    r_squared2 = 1 - (sse2 / sst2)
    print("Коэффициент детерминации (R-squared) 2:", r_squared2)
    # Прогноз на 3 шага вперёд 2 модель
    model2 = LinearRegression()
    model.fit(x_train, y_train2)  # y - входные данные, f_t - фактические значения
    future_input2 = [(193, 181,  140.54365437), (194, 182, 138.4325965435), (195, 183, 142.244325656)]  # Данные для следующих 3 шагов
    forecast2 = model.predict(future_input2)
    print("Прогноз на 3 шага вперед 2 модели:", forecast2)
    print()

    plt.plot(t, f_t, label=f'y(t) = {k1_function} * {a1_function}^{b1_function}^t')
    plt.xlabel('t')
    plt.ylabel('y(t)')
    plt.legend()
    plt.title('График функции y(t)')
    plt.grid(True)
    plt.plot(trend)
    plt.show()